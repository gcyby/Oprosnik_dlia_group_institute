import telebot
import time
import os
from openpyxl.workbook import Workbook
from telebot import types
from openpyxl import load_workbook


TOCKEN ='you tocken'
bot = telebot.TeleBot(TOCKEN)


def user_name(id_user:int,base_name:str,table_name:str,litera_stolbca:str):
    #Находит данные пользователя по первому столбцу и выводит заданное значение
    base = load_workbook(base_name)
    id_group = base[f"{table_name}"]
    name_user = ''
    cnt = 0
    for stroka in id_group['A1:A50']:
        cnt += 1
        for id_klient in stroka:
            if id_klient.value == int(id_user):
                name_user = id_group[f"{litera_stolbca}{cnt}"].value
                break
    return name_user


def time_seckond(format_time):
    format_time = format_time.split(' ')
    year_mounth_day = format_time[0]
    mouth_seconds = int(format_time[0].split('-')[2])*86400
    hour_minute_seconds = format_time[1].split(':')
    time_vote_second = int(hour_minute_seconds[0]) * 3600 + int(hour_minute_seconds[1]) * 60 + int(hour_minute_seconds[2]) + mouth_seconds
    return time_vote_second,year_mounth_day




def sravnim_vremya(masiv_id:list,data_base,callback,id_user:int):
    # Функция отвечает за получение разницы времени и позволяет ли оно переголосовать
    if id_user in masiv_id:
        localtime_second,date_vote = time_seckond(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
        index_user = masiv_id.index(id_user)
        sheet = data_base[f'{callback.message.json.get('text')}']
        year_mounth_day = sheet[f'D{index_user + 1}'].value
        votetime,date_vote2 = time_seckond(year_mounth_day)
        if date_vote == date_vote2 and abs(localtime_second-votetime) <= 1800:
            return True
        else:
            return False
    else:
        return False


def id_user_in_table(file_name:str,table_name:str):
    #Функция получение всех id из листа в эксэле
    id_admin = []
    data_base = load_workbook(f'{file_name}')
    table = data_base[f'{table_name}']
    cnt = 1
    while table[f'A{cnt}'].value is not None:
        id_admin.append(table[f'A{cnt}'].value)
        cnt += 1
    id_admin = [x for x in id_admin if x is not None]
    return id_admin


def sozdaem_opros(danie_oprosa,id_kogoto):
    #Создаёт опрос
    number_group = user_name(id_kogoto,'база старст_групп.xlsx','Group','B')
    id_chat_group = user_name(id_kogoto, 'база старст_групп.xlsx', 'Group', 'C')
    user = danie_oprosa[0]
    danie_oprosa.pop(0)
    colichestvo_voprosov = int(danie_oprosa[0])
    danie_oprosa.pop(0)
    if colichestvo_voprosov == len(danie_oprosa):
        data_base = load_workbook(f'{number_group} база посещений.xlsx')
        new_sheet = data_base.create_sheet(title=user)
        new_sheet["A1"], new_sheet["B1"], new_sheet["C1"], new_sheet["D1"] = 'id', 'имя', 'посещение пары', 'Время ответа'
        data_base.save(f'{number_group} база посещений.xlsx')
        markup = types.InlineKeyboardMarkup()
        for i in range(colichestvo_voprosov):
            markup.add(types.InlineKeyboardButton(danie_oprosa[i], callback_data=f'{number_group} vote{i}'))
        bot.send_message(id_chat_group, user, reply_markup=markup)


if __name__ == '__main__':
    Admin_id = id_user_in_table('база старст_групп.xlsx', 'Admin_List')


    @bot.message_handler(commands = ['start'])
    def start(message):
        if message.from_user.id in Admin_id:
            #Проверяет права доступа к созданию опроса
            markup = types.InlineKeyboardMarkup()
            btn1 = types.InlineKeyboardButton('Сделаем опрос',callback_data='sozdate_opros')
            markup.add(btn1)
            bot.send_message(message.chat.id, 'Привет повелитель хочешь сделать телграм бота', reply_markup=markup)

        else:
            markup = types.InlineKeyboardMarkup()
            btn1 = types.InlineKeyboardButton('Регистрация студента', callback_data='registration_student')
            btn2 = types.InlineKeyboardButton('Регистрация старосты', callback_data='registration_starost')
            markup.add(btn1)
            markup.add(btn2)
            bot.send_message(message.chat.id, 'Привет что-бы продолжить надо зарегистрироваться', reply_markup=markup)


    def registration_starost(message):
        # Функция регистрации старост
        dannie = message.text.split(', ')
        name = dannie[0]
        number_group = dannie[1]
        database = load_workbook('база старст_групп.xlsx')
        sheet = database['Admin_List']
        admin_lst = id_user_in_table('база старст_групп.xlsx', 'Admin_List')
        if message.from_user.id not in admin_lst:
            cnt = 1
            while sheet[f'A{cnt}'].value is not None:
                cnt += 1
            sheet[f'A{cnt}'].value,sheet[f'B{cnt}'].value, sheet[f'C{cnt}'].value = message.from_user.id,name, number_group
            database.save('база старст_групп.xlsx')
            bot.send_message(message.chat.id,'Вы успешно зарегистрированы')
            bot.register_next_step_handler(message,sozdanie_fila_group)


    def registration_student(message):
        # Функция регистрации студента
        dannie = message.text.split(', ')
        masiv_id_group = id_user_in_table(f'{dannie[1]} база посещений.xlsx','id_Group')
        if message.from_user.id not in masiv_id_group:
            database = load_workbook(f'{dannie[1]} база посещений.xlsx')
            sheet = database['id_Group']
            cnt = 1
            while sheet[f'A{cnt}'].value is not None:
                cnt += 1
            sheet[f'A{cnt}'],sheet[f'B{cnt}'],sheet[f'C{cnt}'] = message.from_user.id,message.from_user.username,int(dannie[0])
            database.save(f'{dannie[1]} база посещений.xlsx')
            bot.send_message(message.chat.id,'Спасибо за регистрацию')

        else:
            bot.send_message(message.chat.id,'Вы уже были зарегистрированы')


    def sozdanie_fila_group(message):
        # Создаёт таблицу для группы если её нет при задание старосты
        dannie = message.text.split(', ')
        directory = "C:/Users/user/PycharmProjects/TeleBot"
        file_in_directtory = os.listdir(directory)
        cnt = 0
        for file in file_in_directtory:
            if f'{dannie[1]} база посещений.xlsx' == file:
                cnt+=1
                break
        if cnt == 0:
            data_base = Workbook()
            sheet = data_base.create_sheet('id_Group')
            sheet['A1'].value,sheet['B1'].value,sheet['C1'].value = 'id','user_name','ФИ'
            sheet['A2'].value, sheet['B2'].value, sheet['C2'].value = message.from_user.id,message.from_user.username,dannie[0]
            data_base.save(f'{dannie[1]} база посещений.xlsx')


    def pull_make(message):
        #Функция создания опроса
        qestion=message.text.split(', ')
        id_sozdateli = message.from_user.id
        if len(qestion)>3 and message.from_user.id in Admin_id and qestion[1] in '123456780':
            #Проверяет на соответствие условия создания опроса
            sozdaem_opros(qestion,id_sozdateli)

        else:
            bot.send_message(message.chat.id,'Неправильный ввод данных Пример: "ВышМат-название, 2-колличество вопросов, был, не был"')


    @bot.callback_query_handler(func=lambda callback: True)
    def callback_message(callback):
        id_user = callback.from_user.id
        if 'vote' in callback.data:
            #Заносит в эксэль документ голос человека.
            time_vote = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
            number_group = callback.data.split(' ')[0]
            name_data_base = f'{number_group} база посещений.xlsx'
            data_base = load_workbook(f'{name_data_base}')
            masiv_id = id_user_in_table(name_data_base,callback.message.json.get('text'))
            if id_user not in masiv_id:
                #Проверка голосовал ли человек.
                sheet = data_base[f'{callback.message.json.get('text')}']
                name_user = user_name(id_user,name_data_base,'id_Group','C')
                for i in range(len(callback.message.json.get('reply_markup').get('inline_keyboard'))):
                    if callback.message.json.get('reply_markup').get('inline_keyboard')[i][0].get('callback_data') == callback.data:
                        namber_str = len(masiv_id)+1
                        sheet[f'A{namber_str}'],sheet[f'B{namber_str}'],sheet[f'C{namber_str}'],sheet[f'D{namber_str}'] \
                            = callback.from_user.id,name_user,callback.message.json.get('reply_markup').get('inline_keyboard')[i][0].get('text'),time_vote
                        data_base.save(f'{number_group} база посещений.xlsx')
            elif  (id_user in masiv_id):
                if sravnim_vremya(masiv_id,data_base,callback,id_user) == True:
                    # Проверка голосовал ли человек.
                    sheet = data_base[f'{callback.message.json.get('text')}']
                    name_user = user_name(id_user, name_data_base, 'id_Group', 'C')
                    for i in range(len(callback.message.json.get('reply_markup').get('inline_keyboard'))):
                        if callback.message.json.get('reply_markup').get('inline_keyboard')[i][0].get(
                                'callback_data') == callback.data:
                            namber_str = len(masiv_id)
                            sheet[f'A{namber_str}'], sheet[f'B{namber_str}'], sheet[f'C{namber_str}'], sheet[
                                f'D{namber_str}'] \
                                = callback.from_user.id, name_user, \
                            callback.message.json.get('reply_markup').get('inline_keyboard')[i][0].get('text'), time_vote
                            data_base.save(f'{number_group} база посещений.xlsx')

        elif callback.data == 'sozdate_opros':
            #Создаёт сообщение прокладку для перехода в функцию создания опроса.
            bot.send_message(callback.message.chat.id, 'Назави опрос, колличество ответов и сами ответы через запятую c пробелом')
            bot.register_next_step_handler(callback.message, pull_make)

        elif callback.data == 'registration_student':
            bot.send_message(callback.message.chat.id,'Введи имя фамилию, номер группы писать через пробел с запятой')
            bot.register_next_step_handler(callback.message, registration_student)

        elif callback.data == 'registration_starost':
            bot.send_message(callback.message.chat.id,'Введи имя фамилию, номер группы писать через пробел с запятой')
            bot.register_next_step_handler(callback.message, registration_starost)


    bot.infinity_polling()